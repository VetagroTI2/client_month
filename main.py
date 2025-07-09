#==================Libs Matriz=====================
import pyodbc
import logging
import pandas as pd
import os
import unicodedata
import re
#==================Libs do Excel=====================
from openpyxl.styles import PatternFill, Font, Border, Side
from openpyxl.utils import get_column_letter
#==================Libs do SMTP=====================
import smtplib
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from email.mime.base import MIMEBase
from email import encoders


#configuração dos arquivos de log
log_path = r'C:\EXTRATOR\labyes_sellout.log'
logging.basicConfig(
    filename=log_path,
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s'
)




#   FUNÇÕES
#           DE
#               TRATAMENTO





def normalizar_texto(texto, case='upper', substituir_especiais=False):
    if isinstance(texto, str):
        # Remove acentuação
        nfkd_form = unicodedata.normalize('NFKD', texto)
        texto_sem_acento = nfkd_form.encode('ASCII', 'ignore').decode('utf-8')

        #Remove os espaços extras
        texto_sem_acento = texto_sem_acento.strip()

        #Substituir caracteres especiais (se solicitado)
        if substituir_especiais:
            texto_sem_acento = re.sub(r'[^A-Za-z0-9\s]', '', texto_sem_acento)

        #Ajustar case
        if case == 'upper':
            texto_sem_acento = texto_sem_acento.upper()
        elif case == 'lower':
            texto_sem_acento = texto_sem_acento.lower()
        
        return texto_sem_acento
    return texto

def normalizar_colunas_dataframe(df, colunas=None, case='upper', substituir_especiais=False):
    if colunas is None:
        colunas = df.select_dtypes(include='object').columns.tolist()

    for coluna in colunas:
        if coluna in df.columns:
            df[coluna] = df[coluna].apply(lambda x: normalizar_texto(x, case, substituir_especiais))
        else:
            logging.error(f'Coluna {coluna} não encontrada no DataFrame.')
    return df





#   SCRIPT
#           DE
#               CONSULTA e GRAVAÇÃO




try:
    # Ler o arquivo de conexão principal
    with open(r'C:\EXTRATOR\conexao.txt', 'r') as arquivo:
        next(arquivo)  # Pula a primeira linha
        conexao = ''.join([linha.strip() for linha in arquivo])

    # Estabelece a conexão
    conn = pyodbc.connect(conexao)

    if conn:
        logging.info("Conexao feita!")
    else:
        logging.error("Erro ao estabelecer conexao.")

    # Executa a consulta
    query_movimentacao = '''
        SELECT *
        FROM VW_FATURAMENTO_DETALHADO
        WHERE FOR_CODI = '00262'
        AND LOCALIZACAO = '001'
        AND DATEDIFF(MONTH, CONVERT(DATE, DATA, 103), GETDATE()) = 1
        AND SITUACAO = 'ATIVO'
    '''
    query_estoque = '''
        SELECT * FROM PRODUTO
        INNER JOIN ESTOQUE ON PRODUTO.PRO_CODI = ESTOQUE.PRO_CODI
        WHERE ESTOQUE.LOC_CODI = '001' AND PRODUTO.FOR_CODI = '00262'
    '''
    df_movimentacao = pd.read_sql(query_movimentacao, conn)
    df_estoque = pd.read_sql(query_estoque, conn)

    # Normalizar nomes das colunas
    df_movimentacao.columns = [normalizar_texto(col, case='upper') for col in df_movimentacao.columns]
    df_estoque.columns = [normalizar_texto(col, case='upper') for col in df_estoque.columns]

    # Normalizar os valores das colunas de texto
    df_movimentacao = normalizar_colunas_dataframe(df_movimentacao, case='upper')
    df_estoque = normalizar_colunas_dataframe(df_estoque, case='upper')

    # Verifica se há dados antes de salvar
    if not df_movimentacao.empty and not df_estoque.empty:
        # Filtrar VEN e BOV separadamente no dataframe de movimentação
        df_ven = df_movimentacao[df_movimentacao['NATUREZA'] == 'VEN']
        df_bov = df_movimentacao[df_movimentacao['NATUREZA'] == 'BOV']

        # Filtrar apenas departamento 001 do dataframe de estoque
        df_estoque = df_estoque[df_estoque['DPT_CODI'] == '001']

        # Filtrar apenas as colunas PRO_DESC e EST_QUAN no dataframe de estoque
        df_est_atual = df_estoque[['PRO_DESC', 'EST_QUAN']]

        # Agrupar por PRODUTO e somar QUANTIDADE e VALOR das VENDAS
        df_agrupado_ven = df_ven.groupby('PRODUTO', as_index=False)[['QUANTIDADE', 'VALOR_TOTAL']].sum()
        df_agrupado_ven = df_agrupado_ven.sort_values(by='VALOR_TOTAL', ascending=False)

        # Agrupar por PRODUTO e somar QUANTIDADE e VALOR das BONIFICAÇÕES
        df_agrupado_bov = df_bov.groupby('PRODUTO', as_index=False)[['QUANTIDADE', 'VALOR_TOTAL']].sum()
        df_agrupado_bov = df_agrupado_bov.sort_values(by='VALOR_TOTAL', ascending=False)

        # Renomar os nomes das colunas do df_est_atual
        df_est_atual = df_est_atual.rename(columns={'PRO_DESC':'PRODUTO','EST_QUAN':'QUANTIDADE'})

        caminho_saida = r'C:\EXTRATOR\sellout.xlsx'

        # Salvar ambas as tabelas em abas diferentes
        with pd.ExcelWriter(caminho_saida, engine='openpyxl') as writer:

            # Gravando as informações no excel
            df_agrupado_ven.to_excel(writer, sheet_name='Venda', index=False)
            df_agrupado_bov.to_excel(writer, sheet_name='Bonificação', index=False)
            df_est_atual.to_excel(writer, sheet_name='Estoque Atual', index=False)

            # Cor de fundo vermelha no cabeçalho
            fill = PatternFill(start_color='BF2822', end_color='BF2822', fill_type='solid')
            # Cor da fonte branca e negrito no cabeçalho
            font = Font(color='FFFFFF', bold=True)

            #Aplicar formatação no excel
            abas = {'Venda': df_agrupado_ven, 'Bonificação': df_agrupado_bov, 'Estoque Atual': df_est_atual}

            for sheet_name, df_temp in abas.items():
                worksheet = writer.sheets[sheet_name]

                # Alterar o tipo de dado para moeda brasileira na coluna VALOR_TOTAL
                if sheet_name in ['Venda', 'Bonificação'] and 'VALOR_TOTAL' in df_temp.columns:
                    col_idx = df_temp.columns.get_loc('VALOR_TOTAL') + 1

                    for row in range(2, len(df_temp) + 2):
                        cell = worksheet.cell(row=row, column=col_idx)
                        cell.number_format = 'R$ #.##'

                # Oculta linhas de grade
                worksheet.sheet_view.showGridLines = False

                # Aplicar formatação ao cabeçalho (primeira linha)
                for col_idx, col in enumerate(df_temp.columns, start=1):
                    cell = worksheet.cell(row=1, column=col_idx)
                    cell.fill = fill
                    cell.font = font

                # Calcula largura ideal da coluna com base no nome + conteúdo máximo
                for col_idx, col in enumerate(df_temp.columns, start=1):
                   
                    max_len = max(
                        df_temp[col].astype(str).map(len).max(),
                        len(col)
                    )
                    adjusted_width = max_len + 5  # margem extra
                    worksheet.column_dimensions[get_column_letter(col_idx)].width = adjusted_width
                
                # Aplicar bordar
                borda_fina = Border(
                    left=Side(border_style="thin", color="000000"),
                    right=Side(border_style="thin", color="000000"),
                    top=Side(border_style="thin", color="000000"),
                    bottom=Side(border_style="thin", color="000000")
                )
                for row in worksheet.iter_rows(min_row=1, max_row=len(df_temp) + 1, min_col=1, max_col=len(df_temp.columns)):
                    for cell in row:
                        cell.border = borda_fina

        if os.path.exists(caminho_saida):
            logging.info(f"Arquivo Excel salvo com sucesso em: {caminho_saida}")
        else:
            logging.error(f"Falha ao salvar o arquivo Excel em: {caminho_saida}")
    else:
        logging.warning("DataFrame está vazio. Nenhum dado foi salvo.")

except Exception as e:
    logging.error(f"Erro ao executar consulta: {e}")

finally:
    if 'conn' in locals():
        conn.close()
        logging.info("Conexao com o banco de dados fechada com sucesso.")





#   SCRIPT
#           DE
#               ENVIO DE E-MAIL





# Configurações de envio
smtp_server = 'smtps.uhserver.com'
smtp_port = 465  # SSL usa porta 465

smtp_user = 'ti@veteagro.com.br'
smtp_password = 'Veteagro@16'

from_addr = smtp_user
to_addr = ['ti@veteagro.com.br', 'financeiro@veteagro.com.br', 'vet.agro@uol.com.br']

subject = 'Relatório de Sell-Out - LABYES'
body = '''Prezados,

Segue em anexo o relatório de sell-out do mês anterior.

Atenciosamente,

'''

file_path = r'C:\EXTRATOR\sellout.xlsx'


try:
    # Cria a mensagem
    msg = MIMEMultipart()
    msg['From'] = from_addr
    msg['To'] = ', '.join(to_addr)
    msg['Subject'] = subject

    # Adiciona o corpo do e-mail
    msg.attach(MIMEText(body, 'plain'))

    # Anexa o arquivo
    with open(file_path, 'rb') as attachment:
        part = MIMEBase('application', 'octet-stream')
        part.set_payload(attachment.read())
        encoders.encode_base64(part)
        filename = os.path.basename(file_path)
        part.add_header('Content-Disposition', f'attachment; filename="{filename}"')
        msg.attach(part)

    # Conecta ao servidor SMTP via SSL e envia o e-mail
    with smtplib.SMTP_SSL(smtp_server, smtp_port) as server:
        server.login(smtp_user, smtp_password)
        server.sendmail(from_addr, to_addr, msg.as_string())
        logging.info('E-mail enviado com sucesso.')

except Exception as e:
    logging.error(f'Erro ao enviar o e-mail: {e}')

